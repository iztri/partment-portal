"""
Apartment Detail Portal — Flask Web App
Run:  python app.py
Then: Open http://<your-ip>:5000 on phone/laptop
"""

import json
from io import BytesIO
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
from auth import login_required, role_required, authenticate, logout
from sheets_db import SheetsDB
from config import USERS, MARKETING_CHANNELS, HUB_NAMES

app = Flask(__name__)
app.secret_key = "apartment-portal-secret-key-2024"  # Change in production

@app.template_filter("from_json")
def from_json_filter(value):
    if isinstance(value, str):
        try:
            return json.loads(value)
        except:
            return []
    return value if isinstance(value, list) else []

db = SheetsDB()


# ── Auth Routes ───────────────────────────────────────────────────────────
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        if authenticate(username, password):
            flash(f"Welcome, {session.get('name')}!", "success")
            if session["role"] == "marketing":
                return redirect(url_for("marketing_dashboard"))
            return redirect(url_for("field_dashboard"))
        flash("Invalid username or password", "danger")
    return render_template("login.html")


@app.route("/logout")
def app_logout():
    logout()
    return redirect(url_for("login"))


# ── Marketing Team Routes ─────────────────────────────────────────────────
@app.route("/marketing")
@login_required
@role_required("marketing")
def marketing_dashboard():
    apartments = db.get_all_apartments()
    apartments.reverse()

    # Build latest visit date per apartment
    all_visits = db.get_visits()
    visit_map = {}
    for v in all_visits:
        aid = str(v.get("Apartment ID", ""))
        if aid not in visit_map:
            visit_map[aid] = v
    for apt in apartments:
        aid = str(apt.get("Apartment ID", ""))
        v = visit_map.get(aid)
        apt["_visited_at"] = v["Visited At"] if v else ""
        apt["_visited_by"] = v["Visited By"] if v else ""

    field_users = [
        {"username": u, "name": info["name"]}
        for u, info in USERS.items() if info["role"] == "field"
    ]
    # Standee activity log
    all_assignments = db.get_assignments()
    standee_activity = [
        a for a in all_assignments
        if a.get("placed_at") or a.get("removed_at")
    ][:15]  # latest 15 with timestamps
    # Stats
    total_assignments = len(all_assignments)
    placed_count = sum(1 for a in all_assignments if a.get("status") == "Placed")
    removed_count = sum(1 for a in all_assignments if a.get("status") == "Removed")
    pending_count = sum(1 for a in all_assignments if a.get("status") == "Pending")
    return render_template(
        "marketing_dashboard.html",
        apartments=apartments,
        field_users=field_users,
        marketing_channels=MARKETING_CHANNELS,
        hub_names=HUB_NAMES,
        active="dashboard",
        standee_activity=standee_activity,
        standee_stats={"total": total_assignments, "placed": placed_count, "removed": removed_count, "pending": pending_count},
    )


@app.route("/marketing/add-page")
@login_required
@role_required("marketing")
def marketing_add():
    return render_template("marketing_add.html", hub_names=HUB_NAMES, active="add")


@app.route("/marketing/assign-page")
@login_required
@role_required("marketing")
def marketing_assign():
    all_apts = db.get_all_apartments()
    unassigned = [a for a in all_apts if not a.get("Assigned To")]
    assigned = [a for a in all_apts if a.get("Assigned To")]
    field_users = [
        {"username": u, "name": info["name"]}
        for u, info in USERS.items() if info["role"] == "field"
    ]
    return render_template("marketing_assign.html", unassigned=unassigned,
                           assigned=assigned, field_users=field_users, active="assign")


@app.route("/marketing/reassign", methods=["POST"])
@login_required
@role_required("marketing")
def reassign_apartments():
    apartment_ids = request.form.getlist("reassign_ids[]")
    assigned_to = request.form.get("assigned_to", "").strip()
    assigned_date = request.form.get("assigned_date", "").strip()
    notes_for_field = request.form.get("notes_for_field", "").strip()
    if not apartment_ids or not assigned_to or not assigned_date:
        flash("Select apartments, field user, and date", "danger")
    return redirect(url_for("marketing_assign"))


@app.route("/marketing/field-analysis")
@login_required
@role_required("marketing")
def field_analysis():
    # Gather data from all non-deleted apartments
    all_apts = db.get_all_apartments()
    field_users = [
        {"username": u, "info": info}
        for u, info in USERS.items() if info["role"] == "field"
    ]
    # Aggregate by field user + date
    from collections import defaultdict
    agg = defaultdict(lambda: {"total": 0, "visited": 0, "pending": 0})
    for a in all_apts:
        u = a.get("Assigned To", "")
        d = a.get("Assigned Date", "")
        if not u or not d:
            continue
        key = (u, d)
        agg[key]["total"] += 1
        if a.get("Status") == "Visited":
            agg[key]["visited"] += 1
        else:
            agg[key]["pending"] += 1
    # Sort by date desc then user
    rows = sorted(agg.items(), key=lambda x: (x[0][1], x[0][0]), reverse=True)
    # User totals
    user_totals = defaultdict(lambda: {"total": 0, "visited": 0, "pending": 0})
    for (u, d), s in agg.items():
        user_totals[u]["total"] += s["total"]
        user_totals[u]["visited"] += s["visited"]
        user_totals[u]["pending"] += s["pending"]
    return render_template(
        "marketing_field_analysis.html",
        rows=rows,
        field_users=field_users,
        user_totals=dict(user_totals),
        USERS=USERS,
        active="field_analysis",
    )


@app.route("/marketing/standees")
@login_required
@role_required("marketing")
def marketing_standees():
    standees = db.get_standees()
    all_assignments = db.get_assignments()
    apartments = db.get_all_apartments()
    usage = {s["id"]: db.get_standee_usage(s["id"]) for s in standees}
    # Activity log with timestamps
    standee_activity = [
        a for a in all_assignments
        if a.get("placed_at") or a.get("removed_at")
    ][:20]
    # Stats
    total_assignments = len(all_assignments)
    placed_count = sum(1 for a in all_assignments if a.get("status") == "Placed")
    removed_count = sum(1 for a in all_assignments if a.get("status") == "Removed")
    pending_count = sum(1 for a in all_assignments if a.get("status") == "Pending")
    field_users = [
        {"username": u, "name": info["name"]}
        for u, info in USERS.items() if info["role"] == "field"
    ]
    return render_template(
        "marketing_standees.html",
        standees=standees,
        assignments=all_assignments,
        apartments=apartments,
        usage=usage,
        field_users=field_users,
        active="standees",
        standee_activity=standee_activity,
        standee_stats={"total": total_assignments, "placed": placed_count, "removed": removed_count, "pending": pending_count},
    )


@app.route("/marketing/standees/add", methods=["POST"])
@login_required
@role_required("marketing")
def add_standee():
    name = request.form.get("name", "").strip()
    total = request.form.get("total_units", "0").strip()
    loc = request.form.get("storage_location", "").strip()
    if not name:
        flash("Standee name is required", "danger")
        return redirect(url_for("marketing_standees"))
    result = db.add_standee(name, total, loc)
    if result is None:
        flash(f"Standee '{name}' already exists", "danger")
    else:
        flash(f"Standee '{name}' added ✓", "success")
    return redirect(url_for("marketing_standees"))


@app.route("/marketing/standees/edit/<int:standee_id>", methods=["POST"])
@login_required
@role_required("marketing")
def edit_standee(standee_id):
    name = request.form.get("name", "").strip()
    total = request.form.get("total_units", "0").strip()
    loc = request.form.get("storage_location", "").strip()
    if name:
        db.update_standee(standee_id, name=name, total_units=total, storage_location=loc)
        flash("Standee updated ✓", "success")
    return redirect(url_for("marketing_standees"))


@app.route("/marketing/standees/delete/<int:standee_id>", methods=["POST"])
@login_required
@role_required("marketing")
def delete_standee(standee_id):
    db.delete_standee(standee_id)
    flash("Standee deleted", "success")
    return redirect(url_for("marketing_standees"))


@app.route("/marketing/standees/assign", methods=["POST"])
@login_required
@role_required("marketing")
def assign_standee():
    standee_id = request.form.get("standee_id", "").strip()
    apartment_id = request.form.get("apartment_id", "").strip()
    assigned_to = request.form.get("assigned_to", "").strip()
    start_date = request.form.get("start_date", "").strip()
    end_date = request.form.get("end_date", "").strip()
    quantity = request.form.get("quantity", "0").strip()
    notes = request.form.get("notes", "").strip()
    if not all([standee_id, apartment_id, assigned_to, start_date, end_date, quantity]):
        flash("All fields except notes are required", "danger")
        return redirect(url_for("marketing_standees"))
    db.assign_standee(standee_id, apartment_id, assigned_to, start_date, end_date, quantity, notes)
    flash("Standee assigned ✓", "success")
    return redirect(url_for("marketing_standees"))


@app.route("/marketing/standees/assignments/edit/<int:assignment_id>", methods=["POST"])
@login_required
@role_required("marketing")
def edit_assignment(assignment_id):
    standee_id = request.form.get("standee_id", "").strip()
    apartment_id = request.form.get("apartment_id", "").strip()
    assigned_to = request.form.get("assigned_to", "").strip()
    start_date = request.form.get("start_date", "").strip()
    end_date = request.form.get("end_date", "").strip()
    quantity = request.form.get("quantity", "").strip()
    notes = request.form.get("notes", "").strip()
    kwargs = {}
    if standee_id: kwargs["standee_id"] = int(standee_id)
    if apartment_id: kwargs["apartment_id"] = int(apartment_id)
    if assigned_to: kwargs["assigned_to"] = assigned_to
    if start_date: kwargs["start_date"] = start_date
    if end_date: kwargs["end_date"] = end_date
    if quantity: kwargs["quantity"] = int(quantity)
    kwargs["notes"] = notes
    db.update_assignment(assignment_id, **kwargs)
    flash("Assignment updated ✓", "success")
    return redirect(url_for("marketing_standees"))


@app.route("/field/standees")
@login_required
@role_required("field")
def field_standees():
    username = session.get("user")
    assignments = db.get_assignments(assigned_to=username)
    from datetime import date
    today = date.today().isoformat()
    return render_template(
        "field_standees.html",
        assignments=assignments,
        today=today,
        active='standees',
    )


@app.route("/field/standees/update-status/<int:assignment_id>/<status>", methods=["POST"])
@login_required
@role_required("field")
def update_standee_status(assignment_id, status):
    if status not in ("Placed", "Removed"):
        flash("Invalid status", "danger")
        return redirect(url_for("field_standees"))
    db.update_assignment_status(assignment_id, status)
    flash(f"Standee marked as {status} ✓", "success")
    return redirect(url_for("field_standees"))


@app.route("/marketing/download-page")
@login_required
@role_required("marketing")
def marketing_download():
    return render_template("marketing_download.html", active="download")


@app.route("/marketing/add", methods=["POST"])
@login_required
@role_required("marketing")
def add_apartment():
    name = request.form.get("apartment_name", "").strip()
    hub = request.form.get("hub_name", "").strip()
    link = request.form.get("location_link", "").strip()
    if not name or not hub:
        flash("Apartment name and Hub name are required", "danger")
        return redirect(url_for("marketing_dashboard"))
    db.add_apartment(name, hub, link, session.get("user"))
    flash(f"Apartment '{name}' added!", "success")
    return redirect(url_for("marketing_dashboard"))


@app.route("/marketing/bulk-upload", methods=["POST"])
@login_required
@role_required("marketing")
def bulk_upload():
    data = request.form.get("bulk_data", "").strip()
    if not data:
        flash("No data provided", "danger")
        return redirect(url_for("marketing_dashboard"))

    lines = [l.strip() for l in data.split("\n") if l.strip()]
    added = 0
    errors = []
    for i, line in enumerate(lines, 1):
        parts = [p.strip() for p in line.split(",")]
        if len(parts) < 2 or not parts[0] or not parts[1]:
            errors.append(f"Row {i}: missing name or hub — '{line}'")
            continue
        name, hub = parts[0], parts[1]
        link = parts[2] if len(parts) > 2 else ""
        if hub not in HUB_NAMES:
            errors.append(f"Row {i}: invalid hub '{hub}' — '{name}'")
            continue
        db.add_apartment(name, hub, link, session.get("user"))
        added += 1

    msg = f"✅ {added} apartment(s) added"
    if errors:
        msg += f" | ⚠ {len(errors)} skipped: " + "; ".join(errors[:3])
        if len(errors) > 3:
            msg += f" (+{len(errors) - 3} more)"
    flash(msg, "success" if not errors else "warning")
    return redirect(url_for("marketing_dashboard"))


@app.route("/marketing/delete/<int:apartment_id>", methods=["POST"])
@login_required
@role_required("marketing")
def delete_apartment(apartment_id):
    db.delete_apartment(apartment_id)
    flash(f"Apartment #{apartment_id} moved to trash", "warning")
    return redirect(url_for("marketing_dashboard"))


@app.route("/marketing/trash")
@login_required
@role_required("marketing")
def trash_view():
    if session.get("user") != "gowtham":
        flash("Only Gowtham can access trash", "danger")
        return redirect(url_for("marketing_dashboard"))
    deleted = db.get_deleted_apartments()
    return render_template("trash.html", apartments=deleted, active='trash')


@app.route("/marketing/restore/<int:apartment_id>", methods=["POST"])
@login_required
@role_required("marketing")
def restore_apartment(apartment_id):
    if session.get("user") != "gowtham":
        flash("Only Gowtham can restore apartments", "danger")
        return redirect(url_for("marketing_dashboard"))
    db.restore_apartment(apartment_id)
    flash(f"Apartment #{apartment_id} restored", "success")
    return redirect(url_for("trash_view"))


@app.route("/marketing/download")
@login_required
@role_required("marketing")
def download_data():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()
    status_filter = request.args.get("status", "").strip()

    all_apts = db.get_all_apartments()
    if start_date:
        all_apts = [a for a in all_apts if a.get("Created At", "")[:10] >= start_date]
    if end_date:
        all_apts = [a for a in all_apts if a.get("Created At", "")[:10] <= end_date]
    if status_filter:
        all_apts = [a for a in all_apts if a.get("Status", "") == status_filter]

    all_visits = db.get_visits()
    visit_map = {}
    for v in all_visits:
        aid = str(v.get("Apartment ID", ""))
        if aid not in visit_map:
            visit_map[aid] = v

    wb = Workbook()
    ws = wb.active
    ws.title = "Apartments"

    headers = [
        "Apartment ID", "Apartment Name", "Hub Name", "Location Link",
        "Assigned To", "Assigned Date", "Status", "Created By", "Created At",
        "Visited At", "Visited By", "Manager Name", "Manager Phone", "No of Units", "Notes"
    ]
    ws.append(headers)

    hfill = PatternFill("solid", fgColor="1a237e")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    thin = Border(left=Side("thin","d0d0d0"), right=Side("thin","d0d0d0"),
                  top=Side("thin","d0d0d0"), bottom=Side("thin","d0d0d0"))

    for ci in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=ci)
        c.font = hfont
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin

    for ri, apt in enumerate(all_apts, 2):
        aid = str(apt.get("Apartment ID", ""))
        v = visit_map.get(aid, {})
        vals = [
            apt.get("Apartment ID", ""), apt.get("Apartment Name", ""),
            apt.get("Hub Name", ""), apt.get("Location Link", ""),
            apt.get("Assigned To", ""), apt.get("Assigned Date", ""),
            apt.get("Status", ""), apt.get("Created By", ""), apt.get("Created At", ""),
            v.get("Visited At", "") if v else "",
            v.get("Visited By", "") if v else "",
            v.get("Manager Name", "") if v else "",
            v.get("Manager Phone", "") if v else "",
            v.get("No of Units", "") if v else "",
            v.get("Notes", "") if v else "",
        ]
        ws.append(vals)
        for ci in range(1, len(headers) + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.border = thin

    for ci in range(1, len(headers) + 1):
        mx = len(headers[ci - 1])
        for r in range(2, len(all_apts) + 2):
            v = ws.cell(row=r, column=ci).value
            if v:
                mx = max(mx, len(str(v)))
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(ci)].width = min(mx + 3, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="apartments.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/marketing/assign", methods=["POST"])
@login_required
@role_required("marketing")
def assign_apartments():
    apartment_ids = request.form.getlist("apartment_ids[]")
    assigned_to = request.form.get("assigned_to", "").strip()
    assigned_date = request.form.get("assigned_date", "").strip()
    if not apartment_ids or not assigned_to or not assigned_date:
        flash("Select apartments, field user, and date", "danger")
        return redirect(url_for("marketing_dashboard"))
    db.assign_apartments(apartment_ids, assigned_to, assigned_date)
    flash(f"Assigned {len(apartment_ids)} apartment(s) to {assigned_to}", "success")
    return redirect(url_for("marketing_dashboard"))


@app.route("/marketing/edit/<int:apartment_id>", methods=["POST"])
@login_required
@role_required("marketing")
def edit_apartment(apartment_id):
    name = request.form.get("apartment_name", "").strip()
    hub = request.form.get("hub_name", "").strip()
    link = request.form.get("location_link", "").strip()
    assigned_to = request.form.get("assigned_to", "").strip()
    assigned_date = request.form.get("assigned_date", "").strip()

    updates = {}
    if name:
        updates["Apartment Name"] = name
    if hub:
        updates["Hub Name"] = hub
    if link:
        updates["Location Link"] = link

    # Reassignment resets status to Pending
    if assigned_to or assigned_date:
        # Fetch current record to detect changes
        all_apts = db.get_all_apartments()
        current = next((a for a in all_apts if str(a.get("Apartment ID")) == str(apartment_id)), None)
        old_user = (current or {}).get("Assigned To", "")
        old_date = (current or {}).get("Assigned Date", "")
        if assigned_to and assigned_to != old_user:
            updates["Assigned To"] = assigned_to
        if assigned_date and assigned_date != old_date:
            updates["Assigned Date"] = assigned_date
        if "Assigned To" in updates or "Assigned Date" in updates:
            updates["Status"] = "Pending"
            db.update_apartment(apartment_id, **updates)
            flash(f"Apartment reassigned & reset to Pending ✓", "success")
        else:
            if updates:
                db.update_apartment(apartment_id, **updates)
                flash("Apartment updated!", "success")
            else:
                flash("No changes made", "info")
    else:
        if updates:
            db.update_apartment(apartment_id, **updates)
            flash("Apartment updated!", "success")
        else:
            flash("No changes made", "info")
    return redirect(url_for("marketing_dashboard"))


# ── Field Team Routes ─────────────────────────────────────────────────────
@app.route("/field")
@login_required
@role_required("field")
def field_dashboard():
    username = session.get("user")
    selected_date = request.args.get("date", "")
    apartments = db.get_assigned_for_user(username, selected_date) if selected_date else []
    standee_tasks = db.get_standee_tasks_for_user(username, selected_date) if selected_date else []
    available_dates = db.get_available_dates_for_user(username)
    for apt in apartments:
        apt["_is_revisit"] = bool(apt.get("Notes for Field", "").strip())
    return render_template(
        "field_dashboard.html",
        apartments=apartments,
        standee_tasks=standee_tasks,
        available_dates=available_dates,
        selected_date=selected_date,
        marketing_channels=MARKETING_CHANNELS,
        active='dashboard',
    )


def _channel_key(ch):
    cleaned = ch.replace(' ', '_').replace('/', '_').replace('(', '').replace(')', '').replace("'", '').replace('.', '')
    return f"ch_{cleaned}"

@app.route("/field/visit/<int:apartment_id>")
@login_required
@role_required("field")
def visit_form(apartment_id):
    apartments = db.get_all_apartments()
    apt = None
    for a in apartments:
        if int(a.get("Apartment ID", 0)) == apartment_id:
            apt = a
            break
    if not apt:
        flash("Apartment not found", "danger")
        return redirect(url_for("field_dashboard"))
    visits = db.get_visits(apartment_id)
    latest_visit = visits[-1] if visits else None
    existing_data = {}
    if latest_visit and latest_visit.get("Channels Data (JSON)"):
        try:
            channels_list = json.loads(latest_visit["Channels Data (JSON)"])
            existing_data["channels"] = {c["channel"]: c for c in channels_list}
            existing_data["channels_raw"] = channels_list
        except:
            existing_data["channels"] = {}
            existing_data["channels_raw"] = []
        existing_data["manager_name"] = latest_visit.get("Manager Name", "")
        existing_data["no_of_units"] = latest_visit.get("No of Units", "")
        existing_data["manager_phone"] = latest_visit.get("Manager Phone", "")
        existing_data["notes"] = latest_visit.get("Notes", "")
    return render_template(
        "visit_form.html",
        apartment=apt,
        marketing_channels=MARKETING_CHANNELS,
        existing_data=existing_data,
    )


@app.route("/field/submit-visit", methods=["POST"])
@login_required
@role_required("field")
def submit_visit():
    apartment_id = request.form.get("apartment_id")
    apartment_name = request.form.get("apartment_name", "")
    hub_name = request.form.get("hub_name", "")
    manager_name = request.form.get("manager_name", "").strip()
    no_of_units = request.form.get("no_of_units", "0").strip()
    manager_phone = request.form.get("manager_phone", "").strip()
    notes = request.form.get("notes", "").strip()

    if not manager_name:
        flash("Manager name is required", "danger")
        return redirect(url_for("visit_form", apartment_id=apartment_id))

    # Build channels data from JSON hidden field
    channels_data = []
    channels_json = request.form.get("channels_json", "[]")
    try:
        parsed = json.loads(channels_json)
        if isinstance(parsed, list):
            channels_data = [
                {
                    "channel": c.get("channel", ""),
                    "available": True,
                    "amount": int(c.get("amount", 0) or 0),
                    "days": int(c.get("days", 0) or 0),
                }
                for c in parsed if c.get("channel")
            ]
    except (json.JSONDecodeError, ValueError):
        pass

    db.record_visit(
        apartment_id=apartment_id,
        apartment_name=apartment_name,
        hub_name=hub_name,
        manager_name=manager_name,
        no_of_units=int(no_of_units) if no_of_units else 0,
        manager_phone=manager_phone,
        channels_data=channels_data,
        notes=notes,
        visited_by=session.get("user"),
    )
    flash(f"Visit recorded for {apartment_name}!", "success")
    return redirect(url_for("field_dashboard"))


# ── View Visit Details (both roles) ──────────────────────────────────────
@app.route("/visit/<int:apartment_id>")
@login_required
def view_visit(apartment_id):
    visits = db.get_visits(apartment_id)
    return render_template("view_visit.html", visits=visits, apartment_id=apartment_id)


# ── Home ──────────────────────────────────────────────────────────────────
@app.route("/")
def home():
    if "user" in session:
        if session["role"] == "marketing":
            return redirect(url_for("marketing_dashboard"))
        return redirect(url_for("field_dashboard"))
    return redirect(url_for("login"))


# ── Run ───────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 5000))
    print(f"\n  → Server starting on http://0.0.0.0:{port}")
    print(f"  → Other devices on your network use: http://<YOUR-IP>:{port}")
    print(f"  → Press Ctrl+C to stop\n")
    app.run(host="0.0.0.0", port=port, debug=True)
